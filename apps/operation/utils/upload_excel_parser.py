from datetime import datetime
from openpyxl import load_workbook

EXPECTED_HEADERS = {
    "A": "NUMERO OPERACION",
    "B": "FECHA OPERACION",
    "C": "NOMBRE EMISOR",
    "D": "ID EMISOR",
    "E": "CORREDOR EMISOR",
    "F": "ID CORREDOR EMISOR",
    "G": "NOMBRE PAGADOR",
    "H": "ID PAGADOR",
    "I": "NUMERO FACTURA",
    "J": "ID FACTURA",
    "K": "SALDO FACTURA",
    "L": "BILL FRACTION",
    "M": "NOMBRE INVERSIONISTA",
    "N": "ID INVERSIONISTA",
    "O": "CUENTA INVERSIONISTA",
    "P": "CORREDOR INVERSIONISTA",
    "Q": "FECHA PROBABLE",
    "R": "FECHA FIN",
    "S": "VALOR FUTURO",
    "T": "% DESCUENTO",
    "U": "TASA DESCUENTO",
    "V": "TASA INVERSIONISTA",
}


def normalize_date(value):
    if value in [None, ""]:
        return None

    if hasattr(value, "date"):
        return value.date()

    if isinstance(value, str):
        raw = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue

    return None


def normalize_number(value):
    if value in [None, ""]:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        raw = value.strip().replace("$", "").replace(" ", "")

        # Si viene formato tipo 134.976.505,00
        if "," in raw and "." in raw:
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")

        try:
            return float(raw)
        except ValueError:
            return None

    return None


class UploadExcelParser:
    HEADER_ROW = 6
    DATA_START_ROW = 7

    def parse(self, file_obj):
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active

        self._validate_headers(ws)

        rows = []
        for row_idx in range(self.DATA_START_ROW, ws.max_row + 1):
            current_values = [ws[f"{col}{row_idx}"].value for col in EXPECTED_HEADERS.keys()]
            if all(v in [None, ""] for v in current_values):
                continue

            row = {
                "row_number": row_idx,
                "op_id": ws[f"A{row_idx}"].value,
                "op_date": normalize_date(ws[f"B{row_idx}"].value),
                "emitter_name": ws[f"C{row_idx}"].value,
                "emitter_id": str(ws[f"D{row_idx}"].value).strip() if ws[f"D{row_idx}"].value not in [None, ""] else None,
                "emitter_broker_name": ws[f"E{row_idx}"].value,
                "emitter_broker_id": str(ws[f"F{row_idx}"].value).strip() if ws[f"F{row_idx}"].value not in [None, ""] else None,
                "payer_name": ws[f"G{row_idx}"].value,
                "payer_id": str(ws[f"H{row_idx}"].value).strip() if ws[f"H{row_idx}"].value not in [None, ""] else None,
                "bill_number": ws[f"I{row_idx}"].value,
                "bill_id": str(ws[f"J{row_idx}"].value).strip() if ws[f"J{row_idx}"].value not in [None, ""] else None,
                "bill_balance": normalize_number(ws[f"K{row_idx}"].value),
                "bill_fraction": int(ws[f"L{row_idx}"].value) if ws[f"L{row_idx}"].value not in [None, ""] else None,
                "investor_name": ws[f"M{row_idx}"].value,
                "investor_id": str(ws[f"N{row_idx}"].value).strip() if ws[f"N{row_idx}"].value not in [None, ""] else None,
                "investor_account": str(ws[f"O{row_idx}"].value).strip() if ws[f"O{row_idx}"].value not in [None, ""] else None,
                "investor_broker_name": ws[f"P{row_idx}"].value,
                "fecha_probable": normalize_date(ws[f"Q{row_idx}"].value),
                "fecha_fin": normalize_date(ws[f"R{row_idx}"].value),
                "valor_futuro": normalize_number(ws[f"S{row_idx}"].value),
                "porcentaje_descuento": normalize_number(ws[f"T{row_idx}"].value),
                "tasa_descuento": normalize_number(ws[f"U{row_idx}"].value),
                "tasa_inversionista": normalize_number(ws[f"V{row_idx}"].value),
            }
            rows.append(row)

        return rows

    def _validate_headers(self, ws):
        for col, expected in EXPECTED_HEADERS.items():
            current = ws[f"{col}{self.HEADER_ROW}"].value
            current = str(current).strip() if current is not None else ""
            if current != expected:
                raise ValueError(
                    f"Encabezado inválido en {col}{self.HEADER_ROW}. "
                    f"Esperado '{expected}', recibido '{current}'"
                )