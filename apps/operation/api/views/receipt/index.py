# REST Framework imports
from rest_framework.decorators import APIView
# Models
from apps.operation.models import Receipt
# Serializers
from apps.operation.api.serializers.index import ReceiptSerializer, ReceiptReadOnlySerializer
# Utils
from apps.base.utils.index import response, BaseAV
# Decorators
from apps.base.decorators.index import checkRole
from django.db.models import Q, Count
import logging

from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from math import pow

from django.db import transaction
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework import serializers
import json
from openpyxl import load_workbook
from apps.operation.models import PreOperation, Receipt
from apps.operation.api.serializers.index import ReceiptSerializer
from apps.operation.api.serializers.receipt.index import is_last_active_receipt
from apps.operation.enums import ReceiptStatusEnum
from apps.base.utils.logBalanceAccount import log_balance_change
from drf_spectacular.utils import (
    extend_schema,
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    OpenApiTypes,
    inline_serializer,
)

# Configurar el logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Crear un handler de consola y definir el nivel
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)

# Crear un formato para los mensajes de log
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)

# Añadir el handler al logger
logger.addHandler(console_handler)


# -----------------------------------------------------------------------------
# Schemas auxiliares para drf-spectacular
# -----------------------------------------------------------------------------
# Estos serializers solo documentan entradas/salidas en Swagger. No cambian la
# lógica de negocio actual de recaudos.
ReceiptVoidRequestSchema = inline_serializer(
    name="ReceiptVoidRequest",
    fields={
        "reason": serializers.CharField(required=True, min_length=50, max_length=500, help_text="Motivo obligatorio de anulación. Debe tener entre 50 y 500 caracteres."),
    },
)

ReceiptAdjustRequestSchema = inline_serializer(
    name="ReceiptAdjustRequest",
    fields={
        "reason": serializers.CharField(required=True, min_length=50, max_length=500, help_text="Motivo obligatorio de corrección. Debe tener entre 50 y 500 caracteres."),
        "date": serializers.DateField(required=False, help_text="Nueva fecha de aplicación. Formato: YYYY-MM-DD."),
        "payedAmount": serializers.DecimalField(required=False, max_digits=18, decimal_places=2, help_text="Nuevo monto aplicado."),
        "calculatedDays": serializers.IntegerField(required=False, min_value=0, help_text="Días cálculo. Solo afecta futureValueRecalculation."),
        "receiptStatus": serializers.UUIDField(required=False, help_text="Estado de recaudo. Si no se envía, conserva el actual."),
    },
)

MassiveReceiptPreviewRequestSchema = inline_serializer(
    name="MassiveReceiptPreviewRequest",
    fields={
        "applicationDate": serializers.DateField(required=True, help_text="Fecha de aplicación del recaudo. Formato: YYYY-MM-DD."),
        "receiptStatus": serializers.UUIDField(required=False, allow_null=True),
        "operations": serializers.ListField(
            required=True,
            child=serializers.JSONField(),
            help_text="Lista de operaciones con preOperationId/id, payedAmount y opcionalmente calculatedDays.",
        ),
    },
)

MassiveReceiptUploadExcelRequestSchema = inline_serializer(
    name="MassiveReceiptUploadExcelRequest",
    fields={
        "uploadExcel": serializers.FileField(required=True, help_text="Archivo Excel de recaudos."),
        "context": serializers.CharField(required=True, help_text="JSON string con applicationDate, receiptStatus y rows esperadas."),
    },
)

MassiveReceiptRegisterRequestSchema = inline_serializer(
    name="MassiveReceiptRegisterRequest",
    fields={
        "applicationDate": serializers.DateField(required=True),
        "receiptStatus": serializers.UUIDField(required=True),
        "rows": serializers.ListField(required=False, child=serializers.JSONField()),
        "operations": serializers.ListField(required=False, child=serializers.JSONField()),
    },
)


RECEIPT_REASON_MIN_LENGTH = 50
RECEIPT_REASON_MAX_LENGTH = 500
RECEIPT_REASON_REQUIRED_MESSAGE = "El motivo es obligatorio y debe tener al menos 50 caracteres"


def normalize_receipt_reason(raw_reason):
    """Valida y normaliza motivos de edición/anulación de recaudos."""
    reason = (raw_reason or "").strip()

    if len(reason) < RECEIPT_REASON_MIN_LENGTH:
        raise ValueError(RECEIPT_REASON_REQUIRED_MESSAGE)

    if len(reason) > RECEIPT_REASON_MAX_LENGTH:
        raise ValueError(f"El motivo no puede superar {RECEIPT_REASON_MAX_LENGTH} caracteres.")

    return reason


def receipt_active_queryset():
    """Recaudos vigentes que sí cuentan financieramente."""
    return Receipt.objects.filter(state=1, controlStatus=Receipt.CONTROL_ACTIVE)


def get_user_display_name(user):
    if not user:
        return ""

    full_name = f"{getattr(user, 'first_name', '') or ''} {getattr(user, 'last_name', '') or ''}".strip()
    return full_name or getattr(user, "email", "") or str(getattr(user, "id", "") or "")


def receipt_history_fraction(receipt):
    operation = getattr(receipt, "operation", None)
    bill = getattr(operation, "bill", None) if operation else None

    return (
        getattr(receipt, "fraction", None)
        or getattr(operation, "fraction", None)
        or getattr(operation, "billFraction", None)
        or getattr(bill, "fraction", None)
        or 1
    )


def receipt_history_group_key(receipt):
    operation = getattr(receipt, "operation", None)
    bill = getattr(operation, "bill", None) if operation else None
    fraction = receipt_history_fraction(receipt)

    return f"{getattr(operation, 'id', '')}-{getattr(bill, 'id', '')}-{fraction}"


def build_receipt_change_event(receipt):
    operation = getattr(receipt, "operation", None)
    bill = getattr(operation, "bill", None) if operation else None
    account = getattr(receipt, "account", None) or getattr(operation, "clientAccount", None)
    investor_client = get_account_client(account)
    replaced_by = getattr(receipt, "replacedBy", None)

    is_voided = getattr(receipt, "controlStatus", "") == Receipt.CONTROL_VOIDED
    action_type = "VOIDED" if is_voided else "ADJUSTED"
    action_label = "Anulado" if is_voided else "Editado"
    reason = getattr(receipt, "voidReason", None) if is_voided else getattr(receipt, "adjustmentReason", None)

    event = {
        "id": str(receipt.id),
        "actionType": action_type,
        "actionLabel": action_label,
        "controlStatus": getattr(receipt, "controlStatus", ""),
        "state": getattr(receipt, "state", None),
        "reason": reason or "Sin motivo registrado",
        "voidReason": getattr(receipt, "voidReason", "") or "",
        "adjustmentReason": getattr(receipt, "adjustmentReason", "") or "",
        "date": format_date(getattr(receipt, "date", None)),
        "created_at": format_date(getattr(receipt, "created_at", None)),
        "updated_at": format_date(getattr(receipt, "updated_at", None)),
        "voidedAt": format_date(getattr(receipt, "voidedAt", None)),
        "changedAt": format_date(getattr(receipt, "voidedAt", None) or getattr(receipt, "updated_at", None) or getattr(receipt, "created_at", None)),
        "changedBy": get_user_display_name(getattr(receipt, "voidedBy", None) or getattr(receipt, "user_updated_at", None)),
        "opId": getattr(operation, "opId", None),
        "operationId": str(getattr(operation, "id", "") or ""),
        "billId": getattr(bill, "billId", None),
        "billUniqueId": str(getattr(bill, "id", "") or ""),
        "fraction": int(receipt_history_fraction(receipt) or 1),
        "investorName": get_client_name(investor_client),
        "payerName": getattr(bill, "payerName", None) or "",
        "emitterName": getattr(bill, "emitterName", None) or "",
        "typeReceipt": getattr(getattr(receipt, "typeReceipt", None), "description", None) or "",
        "receiptStatus": getattr(getattr(receipt, "receiptStatus", None), "description", None) or "",
        "payedAmount": float(to_decimal(getattr(receipt, "payedAmount", 0))),
        "realDays": getattr(receipt, "realDays", 0),
        "additionalDays": getattr(receipt, "additionalDays", 0),
        "calculatedDays": getattr(receipt, "calculatedDays", 0),
        "additionalInterests": float(to_decimal(getattr(receipt, "additionalInterests", 0))),
        "investorInterests": float(to_decimal(getattr(receipt, "investorInterests", 0))),
        "presentValueInvestor": float(to_decimal(getattr(receipt, "presentValueInvestor", 0))),
        "remaining": float(to_decimal(getattr(receipt, "remaining", 0))),
        "replacement": None,
    }

    if replaced_by:
        event["replacement"] = {
            "id": str(replaced_by.id),
            "date": format_date(getattr(replaced_by, "date", None)),
            "payedAmount": float(to_decimal(getattr(replaced_by, "payedAmount", 0))),
            "realDays": getattr(replaced_by, "realDays", 0),
            "additionalDays": getattr(replaced_by, "additionalDays", 0),
            "calculatedDays": getattr(replaced_by, "calculatedDays", 0),
            "additionalInterests": float(to_decimal(getattr(replaced_by, "additionalInterests", 0))),
            "presentValueInvestor": float(to_decimal(getattr(replaced_by, "presentValueInvestor", 0))),
            "typeReceipt": getattr(getattr(replaced_by, "typeReceipt", None), "description", None) or "",
            "receiptStatus": getattr(getattr(replaced_by, "receiptStatus", None), "description", None) or "",
            "created_at": format_date(getattr(replaced_by, "created_at", None)),
            "controlStatus": getattr(replaced_by, "controlStatus", ""),
        }

    return event























RECEIPT_TYPE_LABELS = {
    "3d461dea-0545-4a92-a847-31b8327bf033": "Cancelado Anticipado",
    "62b0ca1e-f999-4a76-a07f-be1fe4f38cfb": "Cancelado Vencido",
    "d40e91b1-fb6c-4c61-9da8-78d4f258181d": "Parcial vigente",
    "db1d1fa4-e467-4fde-9aee-bbf4008d688b": "Cancelado Vigente",
    "ed85d2bc-1a4b-45ae-b2fd-f931527d9f7f": "Parcial vencido",
    "edd99cf7-6f47-4c82-a4fd-f13b4c60a0c0": "Parcial anticipado",
}


def get_receipt_label_from_type_id(type_id):
    return RECEIPT_TYPE_LABELS.get(str(type_id), "")

RECEIPT_TYPE_IDS = {
    "CANCELED_ANTICIPATED": "3d461dea-0545-4a92-a847-31b8327bf033",
    "PARTIAL_ANTICIPATED": "edd99cf7-6f47-4c82-a4fd-f13b4c60a0c0",
    "CANCELED_EXPIRED": "62b0ca1e-f999-4a76-a07f-be1fe4f38cfb",
    "PARTIAL_EXPIRED": "ed85d2bc-1a4b-45ae-b2fd-f931527d9f7f",
    "CANCELED_CURRENT": "db1d1fa4-e467-4fde-9aee-bbf4008d688b",
    "PARTIAL_CURRENT": "d40e91b1-fb6c-4c61-9da8-78d4f258181d",
}

STATUS_KEYS = {
    "CANCELED_ANTICIPATED": "canceled_anticipated",
    "PARTIAL_ANTICIPATED": "partial_anticipated",
    "CANCELED_EXPIRED": "canceled_expired",
    "PARTIAL_EXPIRED": "partial_expired",
    "CANCELED_CURRENT": "canceled_current",
    "PARTIAL_CURRENT": "partial_current",
}

SUMMARY_KEYS = {
    "CANCELED_ANTICIPATED": "canceledAnticipated",
    "PARTIAL_ANTICIPATED": "partialAnticipated",
    "CANCELED_EXPIRED": "canceledExpired",
    "PARTIAL_EXPIRED": "partialExpired",
    "CANCELED_CURRENT": "canceledCurrent",
    "PARTIAL_CURRENT": "partialCurrent",
}


def to_decimal(value):
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value))


def money(value):
    return to_decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def parse_date(value):
    if not value:
        return None

    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        if timezone.is_aware(value):
            return timezone.localtime(value).date()
        return value.date()

    if isinstance(value, str):
        clean = value.strip()

        # Caso seguro: YYYY-MM-DD.
        # Este es el formato ideal para fechas de negocio.
        if len(clean) >= 10 and clean[4] == "-" and clean[7] == "-" and "T" not in clean:
            return datetime.strptime(clean[:10], "%Y-%m-%d").date()

        # Caso ISO con hora/zona: 2026-05-16T00:16:00Z
        try:
            iso_value = clean.replace("Z", "+00:00")
            parsed = datetime.fromisoformat(iso_value)

            if timezone.is_aware(parsed):
                return timezone.localtime(parsed).date()

            return parsed.date()
        except Exception:
            pass

        # Último fallback: tomar solo YYYY-MM-DD
        return datetime.strptime(clean[:10], "%Y-%m-%d").date()

    return value


def normalize_to_date(value):
    return parse_date(value)


def format_date(value):
    if not value:
        return None

    if hasattr(value, "isoformat"):
        return value.isoformat()

    return str(value)


def get_client_name(client):
    if not client:
        return ""

    full_name = f"{getattr(client, 'first_name', '') or ''} {getattr(client, 'last_name', '') or ''}".strip()

    return (
        getattr(client, "social_reason", None)
        or getattr(client, "full_name", None)
        or full_name
        or getattr(client, "document_number", None)
        or ""
    )


def get_account_client(account):
    if not account:
        return None

    client = getattr(account, "client", None)

    if client:
        return client

    client_id = getattr(account, "client_id", None)

    if client_id:
        try:
            from apps.client.models import Client
            return Client.objects.filter(id=client_id).first()
        except Exception:
            return None

    return None


def get_account_number(account):
    if not account:
        return ""

    return (
        getattr(account, "account_number", None)
        or getattr(account, "number", None)
        or str(getattr(account, "id", "") or "")
    )


def get_operation_discount_tax(op):
    return (
        getattr(op, "discountTax", None)
        or getattr(op, "discount_tax", None)
        or getattr(op, "tax", None)
        or 0
    )


def get_operation_investor_tax(op):
    return (
        getattr(op, "investorTax", None)
        or getattr(op, "investor_tax", None)
        or 0
    )


def calc_additional_interests(base_amount, tax, days):
    base_amount = float(base_amount or 0)
    tax = float(tax or 0)
    days = int(days or 0)

    if base_amount <= 0 or tax <= 0 or days <= 0:
        return Decimal("0.00")

    grow_factor = pow(1 + (tax / 100), days / 365)
    total = base_amount * grow_factor
    interest = total - base_amount

    return money(interest)


def calc_future_value_recalculation(present_value, investor_tax, calculated_days):
    present_value = float(present_value or 0)
    investor_tax = float(investor_tax or 0)
    calculated_days = int(calculated_days or 0)

    if present_value <= 0 or investor_tax <= 0 or calculated_days <= 0:
        return Decimal("0.00")

    value = present_value * pow(1 + (investor_tax / 100), calculated_days / 365)
    return money(value)


def get_previous_receipt_data(op):
    receipts = Receipt.objects.filter(
        operation_id=op.id,
        state=1,
    ).order_by("-date", "-created_at")

    last_receipt = receipts.first()

    if not last_receipt:
        return {
            "lastDate": None,
            "payedAmount": Decimal("0"),
            "interest": Decimal("0"),
        }

    total_previous_payed = Decimal("0")
    total_previous_interest = Decimal("0")

    for receipt in receipts:
        total_previous_payed += to_decimal(getattr(receipt, "payedAmount", 0))
        total_previous_interest += to_decimal(getattr(receipt, "additionalInterests", 0))

    return {
        "lastDate": getattr(last_receipt, "date", None),
        "payedAmount": total_previous_payed,
        "interest": total_previous_interest,
    }


def get_state(application_date, expiration_date):
    if application_date < expiration_date:
        return "anticipada"

    if application_date > expiration_date:
        return "vencida"

    return "vigente"


def get_receipt_type_key(state_name, canceled):
    if canceled:
        if state_name == "anticipada":
            return "CANCELED_ANTICIPATED"
        if state_name == "vencida":
            return "CANCELED_EXPIRED"
        return "CANCELED_CURRENT"

    if state_name == "anticipada":
        return "PARTIAL_ANTICIPATED"
    if state_name == "vencida":
        return "PARTIAL_EXPIRED"
    return "PARTIAL_CURRENT"

def normalize_to_date(value):
        if not value:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        parsed = parse_date(value)

        if isinstance(parsed, datetime):
            return parsed.date()

        return parsed

def build_receipt_preview_row(op, application_date, receipt_status_id=None, payed_amount_override=None, calculated_days_override=None):
    bill = op.bill
    account = op.clientAccount

    if not bill:
        return None

    investor_client = get_account_client(account)

    op_date = normalize_to_date(getattr(op, "opDate", None))

    expiration_date = normalize_to_date(
        getattr(op, "opExpiration", None)
        or getattr(bill, "expirationDate", None)
        or getattr(op, "expirationDate", None)
        or op_date
    )

    application_date = normalize_to_date(application_date)
    logger.warning(
    "DAYS_DEBUG op_id=%s op_date=%s expiration_date=%s application_date=%s real_days=%s additional_base_days=%s",
    getattr(op, "opId", None),
    op_date,
    expiration_date,
    application_date,
    max((application_date - op_date).days, 0),
    max((application_date - expiration_date).days, 0),
)
    if not op_date:
        op_date = application_date

    if not expiration_date:
        expiration_date = op_date



    current_balance = money(
        getattr(op, "opPendingAmount", None)
        or getattr(bill, "currentBalance", None)
        or getattr(op, "amount", None)
        or 0
    )

    nominal_value = money(
        getattr(op, "payedAmount", None)
        or getattr(bill, "billValue", None)
        or getattr(bill, "total", None)
        or current_balance
    )

    future_value = money(
        getattr(op, "amount", None)
        or current_balance
    )

    present_value_investor_base = money(
        getattr(op, "presentValueInvestor", None)
        or nominal_value
    )

    discount_tax = to_decimal(get_operation_discount_tax(op))
    investor_tax = to_decimal(get_operation_investor_tax(op))

    previous_receipt = get_previous_receipt_data(op)
    last_date = normalize_to_date(previous_receipt.get("lastDate"))

    previous_payed_amount = money(previous_receipt["payedAmount"])
    previous_interest = money(previous_receipt["interest"])

    payed_amount = money(
    payed_amount_override
    if payed_amount_override is not None
    else 0
)

    receipt_status_text = str(receipt_status_id or "")

    state_name = get_state(application_date, expiration_date)

    real_days = max((application_date - op_date).days, 0)

    # Días reales: siempre vienen de fechas.
    # Días cálculo: pueden venir modificados desde el front o desde el Excel.
    # Si no vienen, usamos realDays como valor por defecto para generar la plantilla.
    if calculated_days_override not in [None, ""]:
        try:
            calculated_days = int(calculated_days_override)
        except Exception:
            raise ValueError("Días cálculo inválido.")

        if calculated_days < 0:
            raise ValueError("Días cálculo no puede ser negativo.")
    else:
        calculated_days = real_days

    additional_days = 0
    additional_interests = Decimal("0.00")
    investor_interests = Decimal("0.00")
    additional_interests_sm = Decimal("0.00")
    future_value_recalculation = Decimal("0.00")
    table_remaining = Decimal("0.00")
    remaining = Decimal("0.00")
    present_value_investor = Decimal("0.00")

    # Igual que el individual: primero calculamos intereses y pendiente,
    # luego decidimos si es cancelado o parcial.
    if state_name == "vencida":
        if last_date and previous_interest > 0:
            additional_days = max((application_date - last_date).days, 0)
        else:
            additional_days = max((application_date - expiration_date).days, 0)

        base_for_interest = nominal_value

        if previous_payed_amount > 0:
            base_for_interest = money(
                nominal_value - (previous_payed_amount - previous_interest)
            )

        additional_interests = calc_additional_interests(
            base_for_interest,
            discount_tax,
            additional_days,
        )

        investor_interests = calc_additional_interests(
            base_for_interest,
            investor_tax,
            additional_days,
        )

        additional_interests_sm = money(additional_interests - investor_interests)

    elif state_name == "anticipada":
        additional_days = 0
        additional_interests = Decimal("0.00")
        investor_interests = Decimal("0.00")
        additional_interests_sm = Decimal("0.00")

    else:
        additional_days = 0
        additional_interests = Decimal("0.00")
        investor_interests = Decimal("0.00")
        additional_interests_sm = Decimal("0.00")

    # Pendiente real igual al individual:
    # el pago que amortiza operación es payedAmount - additionalInterests.
    net_for_pending = money(payed_amount - additional_interests)

    if net_for_pending >= current_balance:
        pending_after = Decimal("0.00")
        remaining = money(net_for_pending - current_balance)
        canceled = True
    else:
        pending_after = money(current_balance - net_for_pending)
        remaining = Decimal("0.00")
        canceled = False

    # Si es recaudo de tipo recompra / estado especial, el individual deja remanentes en 0.
    if receipt_status_text == "ea8518e8-168a-46d7-b56a-1286bf0037cd":
        remaining = Decimal("0.00")
        table_remaining = Decimal("0.00")

    # Cálculos dependientes de cancelado/parcial, ya con canceled correcto.
    if state_name == "anticipada":
        if canceled:
            future_value_recalculation = calc_future_value_recalculation(
                present_value_investor_base,
                investor_tax,
                calculated_days,
            )

            if receipt_status_text != "ea8518e8-168a-46d7-b56a-1286bf0037cd":
                table_remaining = money(nominal_value - future_value_recalculation)
            else:
                table_remaining = Decimal("0.00")

            if previous_payed_amount > 0:
                present_value_investor = money(
                    future_value_recalculation - previous_payed_amount
                )
            else:
                present_value_investor = future_value_recalculation
        else:
            future_value_recalculation = Decimal("0.00")
            table_remaining = Decimal("0.00")
            present_value_investor = payed_amount

    elif state_name == "vencida":
        future_value_recalculation = Decimal("0.00")
        table_remaining = Decimal("0.00")

        if canceled:
            if previous_payed_amount > 0:
                present_value_investor = money(
                    payed_amount - additional_interests - remaining
                )
            else:
                present_value_investor = nominal_value
        else:
            present_value_investor = money(payed_amount - additional_interests)

    else:
        future_value_recalculation = Decimal("0.00")
        table_remaining = Decimal("0.00")

        if canceled:
            if previous_payed_amount > 0:
                present_value_investor = money(payed_amount - remaining)
            else:
                present_value_investor = nominal_value
        else:
            present_value_investor = payed_amount

    if present_value_investor < Decimal("0.00"):
        present_value_investor = Decimal("0.00")

    receipt_type_key = get_receipt_type_key(state_name, canceled)
    receipt_type_id = RECEIPT_TYPE_IDS[receipt_type_key]
    status_key = STATUS_KEYS[receipt_type_key]

    to_collect = money(current_balance + additional_interests)

    investor_name = get_client_name(investor_client)
    account_number = get_account_number(account)

    fraction = (
        getattr(op, "fraction", None)
        or getattr(op, "billFraction", None)
        or getattr(bill, "fraction", None)
        or 1
    )

    return {
        "id": str(op.id),
        "preOperationId": str(op.id),
        "operation": str(op.id),

        "billUniqueId": str(bill.id),
        "billId": getattr(bill, "billId", None),
        "fraction": int(fraction or 1),

        "opId": getattr(op, "opId", None),
        "opDate": format_date(op_date),
        "periodStart": format_date(op_date),
        "periodEnd": format_date(expiration_date),

        "emitterId": str(getattr(op, "emitter_id", "") or ""),
        "emitterName": getattr(bill, "emitterName", None) or "",
        "payerId": str(getattr(op, "payer_id", "") or ""),
        "payerName": getattr(bill, "payerName", None) or "",

        "account": str(account.id) if account else None,
        "accountId": str(account.id) if account else None,
        "accountNumber": account_number,
        "investorName": investor_name,

        "futureValue": float(money(future_value)),
        "nominalValue": float(money(nominal_value)),
        "currentBalance": float(money(current_balance)),

        # Para mostrar en paso 2 y Excel:
        # debe ser el opPendingAmount actual, sin intereses adicionales.
        "pending": float(money(current_balance)),

        # Para registrar el recaudo:
        # saldo posterior luego de descontar el monto aplicado.
        "pendingAfter": float(money(pending_after)),

        "payedAmount": float(money(payed_amount)),
        "toCollect": float(money(to_collect)),

        "realDays": real_days,
        "calculatedDays": calculated_days,
        "additionalDays": additional_days,

        "additionalInterests": float(money(additional_interests)),
        "interests": float(money(additional_interests)),
        "investorInterests": float(money(investor_interests)),
        "additionalInterestsSM": float(money(additional_interests_sm)),
        "tableInterests": float(money(additional_interests_sm)),
        "remaining": float(money(remaining)),
        "tableRemaining": float(money(table_remaining)),
        "futureValueRecalculation": float(money(future_value_recalculation)),
        "presentValueInvestor": float(money(present_value_investor)),

        "state": state_name,
        "statusKey": status_key,
        "receiptType": receipt_type_id,
        "typeReceipt": receipt_type_id,
        "receiptStatus": receipt_status_id,

        "previousPayedAmount": float(money(previous_payed_amount)),
        "previousInterest": float(money(previous_interest)),
        "lastDate": format_date(previous_receipt["lastDate"]),
    }

class MassiveReceiptPreview(APIView):
    @extend_schema(
        tags=["Recaudos"],
        summary="Previsualizar recaudos masivos",
        description="Calcula una previsualización de recaudos para operaciones seleccionadas sin registrar movimientos financieros.",
        request=MassiveReceiptPreviewRequestSchema,
        responses={200: OpenApiResponse(description="Previsualización generada"), 400: OpenApiResponse(description="Datos inválidos")},
        examples=[OpenApiExample(
            "Preview básico",
            value={
                "applicationDate": "2026-05-15",
                "receiptStatus": "ea8518e8-168a-46d7-b56a-1286bf0037cd",
                "operations": [{"preOperationId": "uuid-operacion", "payedAmount": 1000, "calculatedDays": 41}],
            },
            request_only=True,
        )],
    )
    def post(self, request):
        application_date = request.data.get("applicationDate")
        receipt_status_id = request.data.get("receiptStatus")
        operations = request.data.get("operations", [])

        if not application_date:
            return Response(
                {"error": True, "message": "Se requiere applicationDate"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not isinstance(operations, list) or len(operations) == 0:
            return Response(
                {"error": True, "message": "Se requieren operaciones"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            application_date = parse_date(application_date)
            logger.warning(
                    "MASSIVE_RECEIPT_PREVIEW application_date parsed=%s raw=%s",
                    application_date,
                    request.data.get("applicationDate"),
                )
        except Exception:
            return Response(
                {"error": True, "message": "applicationDate inválida"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        preoperation_ids = [
            item.get("preOperationId") or item.get("id")
            for item in operations
            if item.get("preOperationId") or item.get("id")
        ]

        payed_by_operation = {
            str(item.get("preOperationId") or item.get("id")): item.get("payedAmount")
            for item in operations
            if item.get("preOperationId") or item.get("id")
        }

        calculated_days_by_operation = {
            str(item.get("preOperationId") or item.get("id")): item.get("calculatedDays")
            for item in operations
            if item.get("preOperationId") or item.get("id")
        }

        preoperations = (
            PreOperation.objects
            .filter(id__in=preoperation_ids, status=1, state=1)
            .select_related("bill", "clientAccount")
        )

        rows = []
        summary = {
            "canceledAnticipated": 0,
            "partialAnticipated": 0,
            "canceledExpired": 0,
            "partialExpired": 0,
            "canceledCurrent": 0,
            "partialCurrent": 0,
        }

        for op in preoperations:
            try:
                row = build_receipt_preview_row(
                    op,
                    application_date,
                    receipt_status_id=receipt_status_id,
                    payed_amount_override=payed_by_operation.get(str(op.id)),
                    calculated_days_override=calculated_days_by_operation.get(str(op.id)),
                )
            except ValueError as error:
                return Response(
                    {
                        "error": True,
                        "message": str(error),
                        "operation": str(op.id),
                        "opId": getattr(op, "opId", None),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if not row:
                continue

            rows.append(row)

            receipt_key = None
            for key, value in STATUS_KEYS.items():
                if value == row["statusKey"]:
                    receipt_key = key
                    break

            if receipt_key and receipt_key in SUMMARY_KEYS:
                summary[SUMMARY_KEYS[receipt_key]] += 1

        return Response(
            {
                "error": False,
                "summary": summary,
                "count": len(rows),
                "data": rows,
            },
            status=status.HTTP_200_OK,
        )


def build_field_errors(errors):
    field_errors = {}

    for error in errors:
        field = error.get("field")
        message = error.get("message")

        if field and message:
            field_errors[field] = message

    return field_errors


def build_receipt_summary_from_rows(rows):
    summary = {
        "canceledAnticipated": 0,
        "partialAnticipated": 0,
        "canceledExpired": 0,
        "partialExpired": 0,
        "canceledCurrent": 0,
        "partialCurrent": 0,
    }

    for row in rows:
        status_key = row.get("statusKey")
        receipt_key = None

        for key, value in STATUS_KEYS.items():
            if value == status_key:
                receipt_key = key
                break

        if receipt_key and receipt_key in SUMMARY_KEYS:
            summary[SUMMARY_KEYS[receipt_key]] += 1

    return summary

class MassiveReceiptUploadExcelParser:
    HEADER_ROW = 6

    HEADER_ALIASES = {
        "FACTURA": ["FACTURA"],
        "FRACCION": ["FRACCIÓN", "FRACCION"],
        "OPID": ["OPID", "OP ID"],
        "PERIODO_INICIO": ["PERIODO INICIO", "FECHA INICIO"],
        "PERIODO_FIN": ["PERIODO FIN", "FECHA FIN"],
        "POR_COBRAR": ["POR COBRAR"],
        "DIAS_CALCULO": ["DÍAS CÁLCULO", "DIAS CALCULO"],
        "MONTO_APLICACION": ["MONTO APLICACIÓN", "MONTO APLICACION"],
        "PREOPERACION": ["PREOPERACIÓN", "PREOPERACION"],
        "ID_FACTURA": ["ID FACTURA"],
    }

    def normalize_header(self, value):
        if value is None:
            return ""

        return (
            str(value)
            .strip()
            .upper()
            .replace("\n", " ")
            .replace("  ", " ")
        )

    def find_col(self, headers, key):
        aliases = self.HEADER_ALIASES.get(key, [])

        for alias in aliases:
            normalized_alias = self.normalize_header(alias)
            if normalized_alias in headers:
                return headers[normalized_alias]

        return None

    def get_cell_value(self, sheet, row_number, col):
        if not col:
            return None

        return sheet.cell(row_number, col).value

    def parse(self, excel_file):
        workbook = load_workbook(excel_file, data_only=True)
        sheet = workbook.active

        headers = {}

        for col in range(1, sheet.max_column + 1):
            value = self.normalize_header(sheet.cell(self.HEADER_ROW, col).value)
            if value:
                headers[value] = col

        required_keys = [
            "FACTURA",
            "FRACCION",
            "OPID",
            "PERIODO_INICIO",
            "PERIODO_FIN",
            "DIAS_CALCULO",
            "MONTO_APLICACION",
            "PREOPERACION",
            "ID_FACTURA",
        ]

        missing = [
            key
            for key in required_keys
            if not self.find_col(headers, key)
        ]

        if missing:
            raise ValueError(
                "El Excel no tiene las columnas requeridas para recaudos: "
                + ", ".join(missing)
            )

        col_factura = self.find_col(headers, "FACTURA")
        col_fraccion = self.find_col(headers, "FRACCION")
        col_opid = self.find_col(headers, "OPID")
        col_inicio = self.find_col(headers, "PERIODO_INICIO")
        col_fin = self.find_col(headers, "PERIODO_FIN")
        col_por_cobrar = self.find_col(headers, "POR_COBRAR")
        col_dias_calculo = self.find_col(headers, "DIAS_CALCULO")
        col_monto = self.find_col(headers, "MONTO_APLICACION")
        col_preoperacion = self.find_col(headers, "PREOPERACION")
        col_id_factura = self.find_col(headers, "ID_FACTURA")

        rows = []

        for row_number in range(self.HEADER_ROW + 1, sheet.max_row + 1):
            bill_number = self.get_cell_value(sheet, row_number, col_factura)
            preoperation_id = self.get_cell_value(sheet, row_number, col_preoperacion)

            if not bill_number and not preoperation_id:
                continue

            rows.append({
                "rowNumber": row_number,
                "billNumber": str(bill_number).strip() if bill_number else "",
                "billId": str(self.get_cell_value(sheet, row_number, col_id_factura) or "").strip(),
                "fraction": self.get_cell_value(sheet, row_number, col_fraccion),
                "opId": self.get_cell_value(sheet, row_number, col_opid),
                "periodStart": self.get_cell_value(sheet, row_number, col_inicio),
                "periodEnd": self.get_cell_value(sheet, row_number, col_fin),
                "toCollect": self.get_cell_value(sheet, row_number, col_por_cobrar),
                "calculatedDays": self.get_cell_value(sheet, row_number, col_dias_calculo),
                "payedAmount": self.get_cell_value(sheet, row_number, col_monto),
                "preOperationId": str(preoperation_id).strip() if preoperation_id else "",
                "operation": str(preoperation_id).strip() if preoperation_id else "",
            })

        return rows


class MassiveReceiptUploadExcel(APIView):
    @extend_schema(
        tags=["Recaudos"],
        summary="Validar Excel de recaudos masivos",
        description="Lee el Excel cargado, valida filas contra el contexto enviado y devuelve filas normalizadas listas para registrar.",
        request=MassiveReceiptUploadExcelRequestSchema,
        responses={200: OpenApiResponse(description="Excel validado"), 400: OpenApiResponse(description="Archivo o contexto inválido")},
    )
    def post(self, request):
        excel_file = request.FILES.get("uploadExcel")

        if not excel_file:
            return Response(
                {"message": "No se recibió el archivo uploadExcel"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        raw_context = request.data.get("context")
        upload_context = None

        if raw_context:
            try:
                upload_context = json.loads(raw_context)
            except Exception:
                return Response(
                    {"message": "El contexto de carga es inválido"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if not upload_context:
            return Response(
                {"message": "Se requiere contexto de carga para validar recaudos"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        application_date = upload_context.get("applicationDate")
        receipt_status_id = upload_context.get("receiptStatus")
        expected_rows = upload_context.get("rows") or []

        if not application_date:
            return Response(
                {"message": "El contexto no contiene Fecha Aplicación"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not receipt_status_id:
            return Response(
                {"message": "El contexto no contiene Estado de Recaudo"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            application_date = parse_date(application_date)
        except Exception:
            return Response(
                {"message": "Fecha Aplicación inválida"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            parser = MassiveReceiptUploadExcelParser()
            parsed_rows = parser.parse(excel_file)
        except ValueError as error:
            return Response(
                {"message": str(error)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as error:
            logger.exception("Error leyendo Excel de recaudos")
            return Response(
                {"message": "No fue posible leer el archivo Excel"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not parsed_rows:
            return Response(
                {"message": "El archivo no contiene filas válidas"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        expected_by_operation = {
            str(item.get("preOperationId") or item.get("operation") or item.get("id")): item
            for item in expected_rows
            if item.get("preOperationId") or item.get("operation") or item.get("id")
        }

        operation_ids = [
            str(row.get("preOperationId") or row.get("operation") or "")
            for row in parsed_rows
            if row.get("preOperationId") or row.get("operation")
        ]

        preoperations = (
            PreOperation.objects
            .filter(id__in=operation_ids, status=1, state=1)
            .select_related("bill", "clientAccount")
        )

        preoperation_map = {
            str(op.id): op
            for op in preoperations
        }

        response_rows = []
        normalized_rows = []
        error_count = 0

        for index, parsed in enumerate(parsed_rows, start=1):
            row_errors = []

            operation_id = str(parsed.get("preOperationId") or parsed.get("operation") or "")
            expected = expected_by_operation.get(operation_id)
            op = preoperation_map.get(operation_id)

            if not operation_id:
                row_errors.append({
                    "field": "preOperationId",
                    "message": "La fila no contiene PREOPERACIÓN.",
                })

            if operation_id and operation_id not in expected_by_operation:
                row_errors.append({
                    "field": "operation_context",
                    "message": "La operación no pertenece al lote generado.",
                })

            if operation_id and not op:
                row_errors.append({
                    "field": "preOperationId",
                    "message": "Operación no encontrada o inactiva.",
                })

            excel_bill_number = str(parsed.get("billNumber") or "").strip()
            expected_bill_number = str((expected or {}).get("billNumber") or "").strip()

            if expected and expected_bill_number and excel_bill_number != expected_bill_number:
                row_errors.append({
                    "field": "billNumber",
                    "message": "La factura no coincide con el lote generado.",
                })

            try:
                payed_amount = money(parsed.get("payedAmount") or 0)
            except Exception:
                payed_amount = Decimal("0.00")
                row_errors.append({
                    "field": "payedAmount",
                    "message": "Monto aplicación inválido.",
                })

            if payed_amount <= Decimal("0.00"):
                row_errors.append({
                    "field": "payedAmount",
                    "message": "El monto aplicación debe ser mayor a cero.",
                })

            try:
                calculated_days = int(parsed.get("calculatedDays") or 0)
            except Exception:
                calculated_days = 0
                row_errors.append({
                    "field": "calculatedDays",
                    "message": "Días cálculo inválido.",
                })

            if calculated_days < 0:
                row_errors.append({
                    "field": "calculatedDays",
                    "message": "Días cálculo no puede ser negativo.",
                })

            preview_row = None

            if op and not row_errors:
                try:
                    preview_row = build_receipt_preview_row(
                        op,
                        application_date,
                        receipt_status_id=receipt_status_id,
                        payed_amount_override=payed_amount,
                        calculated_days_override=calculated_days,
                    )

                    # El usuario puede modificar días cálculo en Excel.
                    # Se calcula y se guarda usando ese valor, no realDays.
                    preview_row["calculatedDays"] = calculated_days

                except ValueError as error:
                    row_errors.append({
                        "field": "operation",
                        "message": str(error),
                    })
                except Exception:
                    logger.exception("Error recalculando recaudo desde Excel")
                    row_errors.append({
                        "field": "operation",
                        "message": "No fue posible recalcular el recaudo.",
                    })

            has_errors = len(row_errors) > 0

            if has_errors:
                error_count += 1

            display_row = {
                "id": parsed.get("rowNumber") or index,
                "rowNumber": parsed.get("rowNumber") or index,
                "preOperationId": operation_id,
                "operation": operation_id,

                "billId": parsed.get("billId") or "",
                "billNumber": excel_bill_number,
                "fraction": parsed.get("fraction") or "",
                "opId": parsed.get("opId") or "",
                "investorName": (
                    preview_row.get("investorName")
                    if preview_row
                    else (expected or {}).get("investorName", "")
                ),

                "periodStart": (
                    preview_row.get("periodStart")
                    if preview_row
                    else parsed.get("periodStart")
                ),
                "periodEnd": (
                    preview_row.get("periodEnd")
                    if preview_row
                    else parsed.get("periodEnd")
                ),

                "futureValue": preview_row.get("futureValue") if preview_row else 0,
                "nominalValue": preview_row.get("nominalValue") if preview_row else 0,
                "pending": preview_row.get("pending") if preview_row else 0,
                "additionalDays": preview_row.get("additionalDays") if preview_row else 0,
                "interests": preview_row.get("interests") if preview_row else 0,
                "toCollect": preview_row.get("toCollect") if preview_row else 0,

                "payedAmount": float(payed_amount),
                "calculatedDays": calculated_days,
                "realDays": preview_row.get("realDays") if preview_row else 0,

                "receiptType": preview_row.get("receiptType") if preview_row else "",
                "typeReceipt": preview_row.get("typeReceipt") if preview_row else "",
                "receiptTypeLabel": (
                    get_receipt_label_from_type_id(preview_row.get("typeReceipt"))
                    if preview_row
                    else ""
                ),
                "receiptStatus": receipt_status_id,
                "statusKey": preview_row.get("statusKey") if preview_row else "",

                "errors": row_errors,
                "fieldErrors": build_field_errors(row_errors),
                "hasErrors": has_errors,
            }

            response_rows.append(display_row)

            if preview_row and not has_errors:
                normalized_rows.append({
                        **preview_row,
                        "rowNumber": parsed.get("rowNumber") or index,
                        "payedAmount": float(payed_amount),
                        "calculatedDays": calculated_days,
                        "receiptTypeLabel": get_receipt_label_from_type_id(preview_row.get("typeReceipt")),
                    })

        can_register = error_count == 0 and len(normalized_rows) > 0

        return Response(
            {
                "success": can_register,
                "canRegister": can_register,
                "totalRows": len(response_rows),
                "processedRows": len(normalized_rows),
                "errorCount": error_count,
                "rows": response_rows,
                "normalizedRows": normalized_rows,
                "summary": build_receipt_summary_from_rows(normalized_rows),
                "message": (
                    f"Se han validado {len(normalized_rows)} recaudo(s)."
                    if can_register
                    else "El Excel contiene errores en los datos. Revise los valores resaltados."
                ),
            },
            status=status.HTTP_200_OK,
        )
class MassiveReceiptRegister(APIView):
    @extend_schema(
        tags=["Recaudos"],
        summary="Registrar recaudos masivos",
        description="Registra varios recaudos de diferentes operaciones. Cada fila se trata por su propia operación, no como lote financiero.",
        request=MassiveReceiptRegisterRequestSchema,
        responses={200: OpenApiResponse(description="Recaudos registrados"), 400: OpenApiResponse(description="Errores de validación")},
    )
    @transaction.atomic
    def post(self, request):
        application_date = request.data.get("applicationDate")
        receipt_status_id = request.data.get("receiptStatus")
        rows = request.data.get("rows") or request.data.get("operations") or []

        if not application_date:
            return Response(
                {"error": True, "message": "Se requiere applicationDate"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not receipt_status_id:
            return Response(
                {"error": True, "message": "Se requiere receiptStatus"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not isinstance(rows, list) or len(rows) == 0:
            return Response(
                {"error": True, "message": "Se requieren filas para registrar"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        application_date = parse_date(application_date)

        preoperation_ids = [
            item.get("preOperationId") or item.get("id") or item.get("operation")
            for item in rows
            if item.get("preOperationId") or item.get("id") or item.get("operation")
        ]

        preoperations = (
            PreOperation.objects
            .filter(id__in=preoperation_ids, status=1, state=1)
            .select_related("bill", "clientAccount")
        )

        preoperation_map = {
            str(op.id): op
            for op in preoperations
        }

        created = []
        errors = []

        for item in rows:
            op_id = str(
                item.get("preOperationId")
                or item.get("id")
                or item.get("operation")
                or ""
            )

            op = preoperation_map.get(op_id)

            if not op:
                errors.append({
                    "operation": op_id,
                    "message": "Operación no encontrada o inactiva",
                })
                continue
            
            try:
                payed_amount = money(item.get("payedAmount") or 0)
            except Exception:
                errors.append({
                    "operation": op_id,
                    "message": "Monto aplicación inválido",
                })
                continue

            if payed_amount <= Decimal("0.00"):
                errors.append({
                    "operation": op_id,
                    "message": "El monto aplicación debe ser mayor a cero",
                })
                continue

            calculated_days_override = None
            if item.get("calculatedDays") not in [None, ""]:
                try:
                    calculated_days_override = int(item.get("calculatedDays"))
                except Exception:
                    errors.append({
                        "operation": op_id,
                        "message": "Días cálculo inválido",
                    })
                    continue

                if calculated_days_override < 0:
                    errors.append({
                        "operation": op_id,
                        "message": "Días cálculo no puede ser negativo",
                    })
                    continue

            preview_row = build_receipt_preview_row(
                op,
                application_date,
                receipt_status_id=receipt_status_id,
                payed_amount_override=payed_amount,
                calculated_days_override=calculated_days_override,
            )
            if not preview_row:
                errors.append({
                    "operation": op_id,
                    "message": "No se pudo calcular el recaudo",
                })
                continue

            if not preview_row.get("account"):
                errors.append({
                    "operation": op_id,
                    "message": "La operación no tiene cuenta de inversionista asociada",
                })
                continue

            payload = {
                "date": application_date,
                "typeReceipt": preview_row["typeReceipt"],
                "receiptStatus": receipt_status_id,

                "payedAmount": preview_row["payedAmount"],
                "opPendingAmount": preview_row["pendingAfter"],
                "operation": preview_row["operation"],
                "account": preview_row["account"],

                "additionalDays": preview_row["additionalDays"],
                "additionalInterests": preview_row["additionalInterests"],
                "additionalInterestsSM": preview_row["additionalInterestsSM"],
                "investorInterests": preview_row["investorInterests"],
                "tableInterests": 0,
                "futureValueRecalculation": preview_row["futureValueRecalculation"],
                "tableRemaining": preview_row["tableRemaining"],
                "realDays": preview_row["realDays"],
                "remaining": preview_row["remaining"],
                "calculatedDays": preview_row["calculatedDays"],
                "presentValueInvestor": preview_row["presentValueInvestor"],

                "client": preview_row.get("payerId") or None,
                "user_created_at": getattr(request.user, "id", None),
            }

            serializer = ReceiptSerializer(
                data=payload,
                context={"request": request},
            )

            if serializer.is_valid():
                receipt = serializer.save()
                created.append(serializer.data)
            else:
                errors.append({
                    "operation": op_id,
                    "billId": preview_row.get("billId"),
                    "message": serializer.errors,
                })

        if errors:
            transaction.set_rollback(True)
            return Response(
                {
                    "error": True,
                    "message": "Algunos recaudos no pudieron registrarse",
                    "errors": errors,
                    "createdCount": len(created),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {
                "error": False,
                "message": "Recaudos registrados correctamente",
                "createdCount": len(created),
                "data": created,
            },
            status=status.HTTP_200_OK,
        )


def get_receipt_balance_movement(receipt):
    return round(float(receipt.presentValueInvestor or 0) + float(receipt.investorInterests or 0), 2)


def restore_receipt_financial_effects(receipt, request_user=None):
    """
    Revierte los efectos financieros de un recaudo.
    Prioridad: usar ReceiptSnapshot. Si no existe snapshot, aplica reversa segura para casos normales.
    """

    operation = receipt.operation
    account = receipt.account
    bill = getattr(operation, "bill", None)

    try:
        snapshot = receipt.snapshot
    except Exception:
        snapshot = None

    if snapshot:
        if account and snapshot.accountBalanceBefore is not None:
            movement = round(float(snapshot.accountBalanceBefore) - float(account.balance or 0), 2)
            if movement != 0:
                log_balance_change(
                    account,
                    account.balance,
                    snapshot.accountBalanceBefore,
                    movement,
                    "receipt_void",
                    receipt.id,
                    "Receipt - void snapshot",
                )
            account.balance = snapshot.accountBalanceBefore
            account.save()

        operation.status = snapshot.operationStatusBefore
        operation.opPendingAmount = snapshot.operationPendingBefore

        if snapshot.operationAmountBefore is not None:
            operation.amount = snapshot.operationAmountBefore

        if snapshot.operationPayedAmountBefore is not None:
            operation.payedAmount = snapshot.operationPayedAmountBefore

        operation.save()

        if bill:
            if snapshot.billCurrentBalanceBefore is not None:
                bill.currentBalance = snapshot.billCurrentBalanceBefore

            if snapshot.billReBuyAvailableBefore is not None:
                bill.reBuyAvailable = snapshot.billReBuyAvailableBefore

            bill.save()

        return

    # Fallback para históricos sin snapshot.
    # Recompra toca factura y estado de operación de forma más compleja: se bloquea sin snapshot.
    if receipt.receiptStatus_id == ReceiptStatusEnum.RECOMPRA.value:
        raise ValueError(
            "Este recaudo histórico no tiene snapshot y es recompra. Debe revisarse manualmente."
        )

    if account:
        movement = get_receipt_balance_movement(receipt)
        new_balance = round(float(account.balance or 0) - movement, 2)
        log_balance_change(
            account,
            account.balance,
            new_balance,
            -movement,
            "receipt_void",
            receipt.id,
            "Receipt - void fallback",
        )
        account.balance = new_balance
        account.save()

    net_paid = round(float(receipt.payedAmount or 0) - float(receipt.additionalInterests or 0), 2)

    if receipt.typeReceipt_id in ReceiptStatusEnum.CANCELED_STATUSES.value:
        restored_pending = max(
            round(net_paid - float(receipt.remaining or 0), 2),
            0,
        )
    else:
        restored_pending = max(
            round(float(operation.opPendingAmount or 0) + net_paid, 2),
            0,
        )

    operation.opPendingAmount = restored_pending
    operation.status = 1 if restored_pending > 0 else 4
    operation.save()


def build_receipt_payload_from_preview(preview_row, application_date, receipt_status_id, extra=None):
    payload = {
        "date": application_date,
        "typeReceipt": preview_row["typeReceipt"],
        "receiptStatus": receipt_status_id,
        "payedAmount": preview_row["payedAmount"],
        "opPendingAmount": preview_row["pendingAfter"],
        "operation": preview_row["operation"],
        "account": preview_row["account"],
        "additionalDays": preview_row["additionalDays"],
        "additionalInterests": preview_row["additionalInterests"],
        "additionalInterestsSM": preview_row["additionalInterestsSM"],
        "investorInterests": preview_row["investorInterests"],
        "tableInterests": 0,
        "futureValueRecalculation": preview_row["futureValueRecalculation"],
        "tableRemaining": preview_row["tableRemaining"],
        "realDays": preview_row["realDays"],
        "remaining": preview_row["remaining"],
        "calculatedDays": preview_row["calculatedDays"],
        "presentValueInvestor": preview_row["presentValueInvestor"],
        "client": preview_row.get("payerId") or None,
    }

    if extra:
        payload.update(extra)

    return payload


class ReceiptVoidAV(APIView):
    @extend_schema(
        tags=["Recaudos"],
        summary="Anular recaudo",
        description="Anula un recaudo solo si es el último recaudo activo de su operación. No borra físicamente; deja trazabilidad y restaura saldos usando snapshot cuando existe.",
        request=ReceiptVoidRequestSchema,
        responses={200: ReceiptReadOnlySerializer, 400: OpenApiResponse(description="No se puede anular o falta motivo"), 404: OpenApiResponse(description="Recaudo no encontrado")},
    )
    @checkRole(['admin'])
    @transaction.atomic
    def post(self, request, pk):
        try:
            receipt = (
                Receipt.objects
                .select_related("operation", "account", "operation__bill", "receiptStatus", "typeReceipt")
                .select_for_update()
                .get(pk=pk, state=1)
            )

            if not is_last_active_receipt(receipt):
                return response({
                    "error": True,
                    "message": "Solo se puede anular el último recaudo activo de la operación.",
                }, 400)

            try:
                reason = normalize_receipt_reason(request.data.get("reason"))
            except ValueError as error:
                return response({"error": True, "message": str(error)}, 400)

            restore_receipt_financial_effects(receipt, request.user)

            receipt.controlStatus = Receipt.CONTROL_VOIDED
            receipt.state = 0
            receipt.voidReason = reason
            receipt.voidedAt = timezone.now()
            receipt.voidedBy = request.user
            receipt.updated_at = timezone.now()
            receipt.user_updated_at = request.user
            receipt.save()

            serializer = ReceiptReadOnlySerializer(receipt)
            return response({
                "error": False,
                "message": "Recaudo anulado correctamente.",
                "data": serializer.data,
            }, 200)
        except Receipt.DoesNotExist:
            return response({"error": True, "message": "Recaudo no encontrado"}, 404)
        except Exception as e:
            transaction.set_rollback(True)
            return response({"error": True, "message": str(e)}, getattr(e, "status_code", 500))


class ReceiptAdjustAV(APIView):
    @extend_schema(
        tags=["Recaudos"],
        summary="Editar/corregir recaudo",
        description="Corrige el último recaudo activo de una operación. Internamente reversa el original, lo marca como ajustado y crea un nuevo recaudo activo corregido.",
        request=ReceiptAdjustRequestSchema,
        responses={200: OpenApiResponse(description="Recaudo corregido"), 400: OpenApiResponse(description="No se puede editar o datos inválidos"), 404: OpenApiResponse(description="Recaudo no encontrado")},
    )
    @checkRole(['admin'])
    @transaction.atomic
    def post(self, request, pk):
        try:
            receipt = (
                Receipt.objects
                .select_related("operation", "account", "operation__bill", "receiptStatus", "typeReceipt")
                .select_for_update()
                .get(pk=pk, state=1)
            )

            if not is_last_active_receipt(receipt):
                return response({
                    "error": True,
                    "message": "Solo se puede editar el último recaudo activo de la operación.",
                }, 400)

            try:
                reason = normalize_receipt_reason(request.data.get("reason"))
            except ValueError as error:
                return response({"error": True, "message": str(error)}, 400)

            application_date = request.data.get("date") or receipt.date
            application_date = parse_date(application_date)
            receipt_status_id = request.data.get("receiptStatus") or receipt.receiptStatus_id

            try:
                payed_amount = money(
                    request.data.get("payedAmount")
                    if request.data.get("payedAmount") not in [None, ""]
                    else receipt.payedAmount
                )
            except Exception:
                return response({"error": True, "message": "Monto aplicación inválido."}, 400)

            if payed_amount <= Decimal("0.00"):
                return response({"error": True, "message": "El monto aplicación debe ser mayor a cero."}, 400)

            calculated_days_override = None
            if request.data.get("calculatedDays") not in [None, ""]:
                try:
                    calculated_days_override = int(request.data.get("calculatedDays"))
                except Exception:
                    return response({"error": True, "message": "Días cálculo inválido."}, 400)

                if calculated_days_override < 0:
                    return response({"error": True, "message": "Días cálculo no puede ser negativo."}, 400)
            else:
                calculated_days_override = receipt.calculatedDays

            # Primero se revierte el movimiento original, para recalcular desde el saldo correcto.
            restore_receipt_financial_effects(receipt, request.user)

            receipt.controlStatus = Receipt.CONTROL_ADJUSTED
            receipt.state = 0
            receipt.adjustmentReason = reason
            receipt.updated_at = timezone.now()
            receipt.user_updated_at = request.user
            receipt.save()

            op = receipt.operation
            preview_row = build_receipt_preview_row(
                op,
                application_date,
                receipt_status_id=receipt_status_id,
                payed_amount_override=payed_amount,
                calculated_days_override=calculated_days_override,
            )

            if not preview_row:
                transaction.set_rollback(True)
                return response({"error": True, "message": "No se pudo recalcular el recaudo."}, 400)

            if not preview_row.get("account"):
                transaction.set_rollback(True)
                return response({"error": True, "message": "La operación no tiene cuenta de inversionista asociada."}, 400)

            payload = build_receipt_payload_from_preview(
                preview_row,
                application_date,
                receipt_status_id,
                extra={
                    "originalReceipt": receipt.id,
                    "adjustmentReason": reason,
                    "controlStatus": Receipt.CONTROL_ACTIVE,
                    "user_created_at": getattr(request.user, "id", None),
                },
            )

            serializer = ReceiptSerializer(data=payload, context={"request": request})

            if serializer.is_valid():
                new_receipt = serializer.save()
                receipt.replacedBy = new_receipt
                receipt.save(update_fields=["replacedBy"])
                return response({
                    "error": False,
                    "message": "Recaudo ajustado correctamente.",
                    "data": ReceiptReadOnlySerializer(new_receipt).data,
                    "original": ReceiptReadOnlySerializer(receipt).data,
                }, 200)

            transaction.set_rollback(True)
            return response({"error": True, "message": serializer.errors}, 400)
        except Receipt.DoesNotExist:
            return response({"error": True, "message": "Recaudo no encontrado"}, 404)
        except Exception as e:
            transaction.set_rollback(True)
            return response({"error": True, "message": str(e)}, getattr(e, "status_code", 500))


class ReceiptHistoryAV(BaseAV):
    @extend_schema(
        tags=["Recaudos"],
        summary="Historial de cambios de recaudos",
        description="Agrupa en backend las ediciones/anulaciones de recaudos por operación, factura y fracción. El frontend solo renderiza los grupos devueltos.",
        parameters=[
            OpenApiParameter("page", OpenApiTypes.INT, OpenApiParameter.QUERY, required=False),
            OpenApiParameter("pageSize", OpenApiTypes.INT, OpenApiParameter.QUERY, required=False),
            OpenApiParameter("opId_billId", OpenApiTypes.STR, OpenApiParameter.QUERY, required=False, description="Buscar por opID o factura"),
            OpenApiParameter("startDate", OpenApiTypes.DATE, OpenApiParameter.QUERY, required=False, description="Filtra por fecha de movimiento desde"),
            OpenApiParameter("endDate", OpenApiTypes.DATE, OpenApiParameter.QUERY, required=False, description="Filtra por fecha de movimiento hasta"),
        ],
        responses={200: OpenApiResponse(description="Grupos de historial de cambios")},
    )
    @checkRole(['admin'])
    def get(self, request):
        try:
            search_value = (request.query_params.get("opId_billId") or request.query_params.get("search") or "").strip()
            start_date = request.query_params.get("startDate")
            end_date = request.query_params.get("endDate")
            page = int(request.query_params.get("page") or 1)
            page_size = int(request.query_params.get("pageSize") or request.query_params.get("page_size") or 15)

            if page < 1:
                page = 1

            if page_size < 1:
                page_size = 15

            queryset = (
                Receipt.objects
                .filter(
                    Q(controlStatus__in=[Receipt.CONTROL_VOIDED, Receipt.CONTROL_ADJUSTED])
                    | Q(voidReason__isnull=False)
                    | Q(adjustmentReason__isnull=False)
                )
                .select_related(
                    "operation",
                    "operation__bill",
                    "account",
                    "account__client",
                    "typeReceipt",
                    "receiptStatus",
                    "voidedBy",
                    "user_updated_at",
                    "originalReceipt",
                    "replacedBy",
                    "replacedBy__typeReceipt",
                    "replacedBy__receiptStatus",
                )
                .order_by("-updated_at", "-voidedAt", "-created_at")
            )

            if search_value:
                search_filter = Q(operation__bill__billId__icontains=search_value)
                try:
                    search_filter |= Q(operation__opId=int(search_value))
                except (ValueError, TypeError):
                    pass
                queryset = queryset.filter(search_filter)

            if start_date not in [None, ""]:
                queryset = queryset.filter(
                    Q(updated_at__date__gte=start_date)
                    | Q(voidedAt__date__gte=start_date)
                    | Q(created_at__date__gte=start_date)
                )

            if end_date not in [None, ""]:
                queryset = queryset.filter(
                    Q(updated_at__date__lte=end_date)
                    | Q(voidedAt__date__lte=end_date)
                    | Q(created_at__date__lte=end_date)
                )

            groups_map = {}

            for receipt in queryset:
                operation = getattr(receipt, "operation", None)
                bill = getattr(operation, "bill", None) if operation else None
                account = getattr(receipt, "account", None) or getattr(operation, "clientAccount", None)
                investor_client = get_account_client(account)
                fraction = int(receipt_history_fraction(receipt) or 1)
                key = receipt_history_group_key(receipt)

                if key not in groups_map:
                    groups_map[key] = {
                        "id": key,
                        "operationId": str(getattr(operation, "id", "") or ""),
                        "opId": getattr(operation, "opId", None),
                        "billUniqueId": str(getattr(bill, "id", "") or ""),
                        "billId": getattr(bill, "billId", None) or "",
                        "fraction": fraction,
                        "investorName": get_client_name(investor_client),
                        "payerName": getattr(bill, "payerName", None) or "",
                        "emitterName": getattr(bill, "emitterName", None) or "",
                        "receipts": [],
                    }

                groups_map[key]["receipts"].append(build_receipt_change_event(receipt))

            groups = []
            for group in groups_map.values():
                group["receipts"] = sorted(
                    group["receipts"],
                    key=lambda item: item.get("changedAt") or item.get("updated_at") or item.get("created_at") or "",
                    reverse=True,
                )
                group["changesCount"] = len(group["receipts"])
                group["lastChange"] = group["receipts"][0] if group["receipts"] else None
                groups.append(group)

            groups = sorted(
                groups,
                key=lambda item: (item.get("lastChange") or {}).get("changedAt") or "",
                reverse=True,
            )

            total = len(groups)
            start = (page - 1) * page_size
            end = start + page_size
            page_groups = groups[start:end]

            return Response(
                {
                    "error": False,
                    "count": total,
                    "page": page,
                    "pageSize": page_size,
                    "totalPages": (total + page_size - 1) // page_size if page_size else 1,
                    "data": page_groups,
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return response({"error": True, "message": str(e)}, getattr(e, "status_code", 500))


class ReceiptAV(BaseAV):

    @extend_schema(
        tags=["Recaudos"],
        summary="Listar o consultar recaudos vigentes",
        description="Devuelve únicamente recaudos activos/vigentes: state=1 y controlStatus=ACTIVE. Los recaudos anulados o ajustados se consultan por /receipt/history/.",
        parameters=[
            OpenApiParameter("id", OpenApiTypes.UUID, OpenApiParameter.QUERY, required=False),
            OpenApiParameter("opId", OpenApiTypes.STR, OpenApiParameter.QUERY, required=False),
            OpenApiParameter("opId_billId", OpenApiTypes.STR, OpenApiParameter.QUERY, required=False),
            OpenApiParameter("statusReceipt", OpenApiTypes.UUID, OpenApiParameter.QUERY, required=False),
            OpenApiParameter("startDate", OpenApiTypes.DATE, OpenApiParameter.QUERY, required=False),
            OpenApiParameter("endDate", OpenApiTypes.DATE, OpenApiParameter.QUERY, required=False),
        ],
        responses={200: ReceiptReadOnlySerializer},
    )
    @checkRole(['admin'])
    def get(self, request, pk=None):
        try:
            base_receipts = (
                receipt_active_queryset()
                .select_related("operation", "operation__bill", "account", "account__client", "typeReceipt", "receiptStatus", "originalReceipt", "replacedBy")
                .order_by("-created_at")
            )

            if pk:
                receipt = base_receipts.get(pk=pk)
                serializer = ReceiptReadOnlySerializer(receipt)
                return response({'error': False, 'data': serializer.data}, 200)

            receipts = base_receipts

            receipt_id = request.query_params.get('id')
            op_id = request.query_params.get('opId')
            search_value = (request.query_params.get('opId_billId') or '').strip()
            status_receipt = request.query_params.get('statusReceipt')
            start_date = request.query_params.get('startDate')
            end_date = request.query_params.get('endDate')

            if receipt_id not in [None, '']:
                receipts = receipts.filter(id=receipt_id)

            if op_id not in [None, '']:
                receipts = receipts.filter(operation__opId=op_id)

            if search_value:
                search_filter = Q(operation__bill__billId__icontains=search_value)
                try:
                    search_filter |= Q(operation__opId=int(search_value))
                except (ValueError, TypeError):
                    pass
                receipts = receipts.filter(search_filter)

            if status_receipt not in [None, '']:
                receipts = receipts.filter(typeReceipt_id=status_receipt)

            if start_date not in [None, '']:
                receipts = receipts.filter(date__gte=start_date)

            if end_date not in [None, '']:
                receipts = receipts.filter(date__lte=end_date)

            page = self.paginate_queryset(receipts)
            if page is not None:
                serializer = ReceiptReadOnlySerializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = ReceiptReadOnlySerializer(receipts, many=True)
            return response({'error': False, 'data': serializer.data}, 200)
        except Receipt.DoesNotExist:
            return response({'error': True, 'message': 'recaudo/s no encontrados'}, 404)
        except Exception as e:
            return response({'error': True, 'message': str(e)}, e.status_code if hasattr(e, 'status_code') else 500)

    @extend_schema(
        tags=["Recaudos"],
        summary="Crear recaudo individual",
        request=ReceiptSerializer,
        responses={200: ReceiptSerializer, 400: OpenApiResponse(description="Errores de validación")},
    )
    @checkRole(['admin'])
    def post(self, request):
        try:
            serializer = ReceiptSerializer(data=request.data, context={'request': request})
            if serializer.is_valid():
                serializer.save()
                return response({'error': False, 'message': 'Recaudo creado','data': serializer.data}, 200)
            else:
                return response({'error': True, 'message': serializer.errors}, 400)
        except Exception as e:
            return response({'error': True, 'message': str(e)}, e.status_code if hasattr(e, 'status_code') else 500)

    @checkRole(['admin'])
    def patch(self, request, pk):
        try:
            receipt     = Receipt.objects.get(pk=pk)
            serializer  = ReceiptSerializer(receipt, data=request.data, context={'request': request}, partial=True)
            if serializer.is_valid():
                serializer.save()
                return response({'error': False, 'message': 'Recaudo actualizado', 'data': serializer.data}, 200)
            return response({'error': True, 'message': serializer.errors}, 400)
        except Receipt.DoesNotExist:
            return response({'error': True, 'message': 'Recaudo no encontrado'}, 404)
        except Exception as e:
            return response({'error': True, 'message': str(e)}, e.status_code if hasattr(e, 'status_code') else 500)
    
    @checkRole(['admin'])
    @transaction.atomic
    def delete(self, request, pk):
        # Compatibilidad temporal: DELETE usa el nuevo flujo de anulación controlada.
        try:
            receipt = (
                Receipt.objects
                .select_related("operation", "account", "operation__bill", "receiptStatus", "typeReceipt")
                .select_for_update()
                .get(pk=pk, state=1)
            )

            if not is_last_active_receipt(receipt):
                return response({
                    "error": True,
                    "message": "Solo se puede anular el último recaudo activo de la operación.",
                }, 400)

            try:
                reason = normalize_receipt_reason(request.data.get("reason"))
            except ValueError as error:
                return response({"error": True, "message": str(error)}, 400)

            restore_receipt_financial_effects(receipt, request.user)

            receipt.controlStatus = Receipt.CONTROL_VOIDED
            receipt.state = 0
            receipt.voidReason = reason
            receipt.voidedAt = timezone.now()
            receipt.voidedBy = request.user
            receipt.updated_at = timezone.now()
            receipt.user_updated_at = request.user
            receipt.save()

            return response({'error': False, 'message': 'Recaudo anulado correctamente'}, 200)
        except Receipt.DoesNotExist:
            return response({'error': True, 'message': 'Recaudo no encontrado'}, 404)
        except Exception as e:
            transaction.set_rollback(True)
            return response({'error': True, 'message': str(e)}, e.status_code if hasattr(e, 'status_code') else 500)


